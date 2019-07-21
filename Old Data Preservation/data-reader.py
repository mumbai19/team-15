from openpyxl import load_workbook

filepath="products.xlsx"

wb=load_workbook(filepath)

sheet=wb.active

max_row=sheet.max_row

max_column=sheet.max_column
sql = []
count = 0
for i in range(2,max_row+1):
	temp = ""
	temp += "INSERT INTO `manualdata`(`product`, `name`, `quantity`, `total`) VALUES ("
	count = 0
	for j in range(1,max_column+1):
		cell_obj=sheet.cell(row=i,column=j)
		if count!=0:
			temp += ","
		count += 1
		temp += "'"+str(cell_obj.value)+"'"
	temp += ");"	
	sql.append(temp)

import mysql.connector



print(sql)
for x in sql:
	mydb = mysql.connector.connect(
	  host="localhost",
	  user="root",
	  passwd="root",
	  database="trishul"
	)

	mycursor = mydb.cursor()
	mycursor.execute(sql)
	mydb.commit()
	mycursor.close()
	mydb.close()

print(mycursor.rowcount, "record inserted.")